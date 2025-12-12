#!/usr/bin/env python3
"""
Script for downloading files from URLs found in XLS, XLSX, or CSV files.
Reads all cells in a file and downloads any external file links to a specified folder.
"""

import argparse
import csv
import os
import re
import sys
from pathlib import Path
from typing import List, Tuple
from urllib.parse import urlparse, unquote
from urllib.request import urlopen, Request
from urllib.error import URLError, HTTPError

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


# URL pattern to detect external links
URL_PATTERN = re.compile(
    r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'
)


def extract_urls_from_text(text: str) -> List[str]:
    """
    Extract all URLs from a text string.

    Args:
        text: Text to search for URLs

    Returns:
        List of URLs found in the text
    """
    if not isinstance(text, str):
        text = str(text)

    urls = URL_PATTERN.findall(text)
    return urls


def get_filename_from_url(url: str) -> str:
    """
    Extract filename from URL. If no filename found, generate one.

    Args:
        url: The URL to extract filename from

    Returns:
        Filename string
    """
    parsed_url = urlparse(url)
    path = unquote(parsed_url.path)

    # Get the last part of the path
    filename = os.path.basename(path)

    # If no filename or it's a directory, generate from URL
    if not filename or '.' not in filename:
        # Use a hash of the URL to generate a unique filename
        url_hash = str(hash(url))[-8:]
        filename = f"downloaded_file_{url_hash}"

    return filename


def download_file(url: str, output_path: Path) -> bool:
    """
    Download a file from URL to the specified path.

    Args:
        url: URL to download from
        output_path: Path where to save the file

    Returns:
        True if successful, False otherwise
    """
    try:
        # Create a request with a user agent to avoid blocking
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        request = Request(url, headers=headers)

        # Download the file
        with urlopen(request, timeout=30) as response:
            # Read the file content
            content = response.read()

            # Save to file
            with open(output_path, 'wb') as f:
                f.write(content)

        return True

    except (URLError, HTTPError, TimeoutError) as e:
        print(f"Error downloading {url}: {str(e)}")
        return False
    except Exception as e:
        print(f"Unexpected error downloading {url}: {str(e)}")
        return False


def read_csv_file(file_path: Path, column_index_name: int = None) -> List[Tuple[str, str]]:
    """
    Read CSV file and extract all URLs from cells.

    Args:
        file_path: Path to CSV file
        column_index_name: Optional column index for custom filename (0-based)

    Returns:
        List of tuples (URL, custom_filename) found in the file
    """
    url_data = []

    try:
        with open(file_path, 'r', encoding='utf-8', newline='') as f:
            reader = csv.reader(f)
            for row in reader:
                custom_name = ""
                if column_index_name is not None and column_index_name < len(row):
                    custom_name = row[column_index_name].strip() if row[column_index_name] else ""

                for cell in row:
                    if cell:
                        found_urls = extract_urls_from_text(cell)
                        for url in found_urls:
                            url_data.append((url, custom_name))
    except UnicodeDecodeError:
        # Try with different encoding
        try:
            with open(file_path, 'r', encoding='cp1251', newline='') as f:
                reader = csv.reader(f)
                for row in reader:
                    custom_name = ""
                    if column_index_name is not None and column_index_name < len(row):
                        custom_name = row[column_index_name].strip() if row[column_index_name] else ""

                    for cell in row:
                        if cell:
                            found_urls = extract_urls_from_text(cell)
                            for url in found_urls:
                                url_data.append((url, custom_name))
        except Exception as e:
            print(f"Error reading CSV file {file_path}: {str(e)}")
    except Exception as e:
        print(f"Error reading CSV file {file_path}: {str(e)}")

    return url_data


def read_xlsx_file(file_path: Path, column_index_name: int = None) -> List[Tuple[str, str]]:
    """
    Read XLSX file and extract all URLs from cells.

    Args:
        file_path: Path to XLSX file
        column_index_name: Optional column index for custom filename (0-based)

    Returns:
        List of tuples (URL, custom_filename) found in the file
    """
    if not OPENPYXL_AVAILABLE:
        print("Error: openpyxl library is not installed. Install it with: pip install openpyxl")
        return []

    url_data = []

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        # Iterate through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Iterate through all rows
            for row in sheet.iter_rows():
                custom_name = ""
                if column_index_name is not None and column_index_name < len(row):
                    cell_value = row[column_index_name].value
                    custom_name = str(cell_value).strip() if cell_value else ""

                for cell in row:
                    if cell.value:
                        found_urls = []
                        # Check if cell has a hyperlink
                        if cell.hyperlink and cell.hyperlink.target:
                            found_urls.append(cell.hyperlink.target)

                        # Also check cell value for URLs
                        found_urls.extend(extract_urls_from_text(str(cell.value)))

                        for url in found_urls:
                            url_data.append((url, custom_name))

        workbook.close()
    except Exception as e:
        print(f"Error reading XLSX file {file_path}: {str(e)}")

    return url_data


def read_xls_file(file_path: Path, column_index_name: int = None) -> List[Tuple[str, str]]:
    """
    Read XLS file and extract all URLs from cells.

    Args:
        file_path: Path to XLS file
        column_index_name: Optional column index for custom filename (0-based)

    Returns:
        List of tuples (URL, custom_filename) found in the file
    """
    if not XLRD_AVAILABLE:
        print("Error: xlrd library is not installed. Install it with: pip install xlrd")
        return []

    url_data = []

    try:
        workbook = xlrd.open_workbook(file_path)

        # Iterate through all sheets
        for sheet_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_index)

            # Iterate through all rows
            for row_index in range(sheet.nrows):
                custom_name = ""
                if column_index_name is not None and column_index_name < sheet.ncols:
                    cell_value = sheet.cell(row_index, column_index_name).value
                    custom_name = str(cell_value).strip() if cell_value else ""

                for col_index in range(sheet.ncols):
                    cell = sheet.cell(row_index, col_index)
                    if cell.value:
                        found_urls = extract_urls_from_text(str(cell.value))
                        for url in found_urls:
                            url_data.append((url, custom_name))

        # Also check for hyperlinks
        try:
            for sheet_index in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_index)
                if hasattr(sheet, 'hyperlink_map'):
                    for row_index, link in sheet.hyperlink_map.items():
                        if link.url_or_path:
                            custom_name = ""
                            if column_index_name is not None and column_index_name < sheet.ncols:
                                cell_value = sheet.cell(row_index, column_index_name).value
                                custom_name = str(cell_value).strip() if cell_value else ""
                            url_data.append((link.url_or_path, custom_name))
        except:
            pass  # Hyperlinks not available in this version

    except Exception as e:
        print(f"Error reading XLS file {file_path}: {str(e)}")

    return url_data


def read_file(file_path: Path, column_index_name: int = None) -> List[Tuple[str, str]]:
    """
    Read file and extract URLs based on file extension.

    Args:
        file_path: Path to the file
        column_index_name: Optional column index for custom filename (0-based)

    Returns:
        List of tuples (URL, custom_filename) found in the file
    """
    suffix = file_path.suffix.lower()

    if suffix == '.csv':
        return read_csv_file(file_path, column_index_name)
    elif suffix == '.xlsx':
        return read_xlsx_file(file_path, column_index_name)
    elif suffix == '.xls':
        return read_xls_file(file_path, column_index_name)
    else:
        print(f"Error: Unsupported file format '{suffix}'. Supported formats: .xls, .xlsx, .csv")
        return []


def main():
    """Main entry point for the script."""
    parser = argparse.ArgumentParser(
        description="Download files from URLs found in XLS, XLSX, or CSV files.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python download.py data.xlsx /path/to/output
  python download.py data.csv ~/Downloads
  python download.py data.xls ./files
        """
    )

    parser.add_argument(
        "file",
        type=str,
        help="Path to XLS, XLSX, or CSV file containing URLs"
    )

    parser.add_argument(
        "folder",
        type=str,
        help="Path to folder where to download files"
    )

    parser.add_argument(
        "--column-index-name",
        type=int,
        default=None,
        help="Column index (0-based) for custom filename. If specified, use value from this column as filename"
    )

    args = parser.parse_args()

    # Convert paths to Path objects
    file_path = Path(args.file).resolve()
    folder_path = Path(args.folder).resolve()

    # Check if input file exists
    if not file_path.exists():
        print(f"Error: File '{file_path}' does not exist.")
        sys.exit(1)

    if not file_path.is_file():
        print(f"Error: '{file_path}' is not a file.")
        sys.exit(1)

    # Create output folder if it doesn't exist
    folder_path.mkdir(parents=True, exist_ok=True)
    print(f"Output directory: {folder_path}")

    # Read file and extract URLs with custom filenames
    print(f"Reading file: {file_path}")
    url_data = read_file(file_path, args.column_index_name)

    # Remove duplicates while preserving order
    seen = set()
    unique_url_data = []
    for url, custom_name in url_data:
        if url not in seen:
            seen.add(url)
            unique_url_data.append((url, custom_name))

    if not unique_url_data:
        print("No URLs found in the file.")
        sys.exit(0)

    print(f"Found {len(unique_url_data)} unique URL(s) in the file")

    # Download each file
    successful = 0
    failed = 0
    skipped = 0
    renamed = 0

    for i, (url, custom_name) in enumerate(unique_url_data, 1):
        print(f"\n[{i}/{len(unique_url_data)}] Processing: {url}")

        # Get filename from URL
        original_filename = get_filename_from_url(url)
        original_path = folder_path / original_filename

        # Determine final filename
        final_filename = original_filename
        final_path = original_path

        if custom_name:
            # Preserve file extension from original filename
            original_ext = Path(original_filename).suffix
            # If custom_name already has extension, use it as is, otherwise add original extension
            if Path(custom_name).suffix:
                final_filename = custom_name
            else:
                final_filename = custom_name + original_ext
            final_path = folder_path / final_filename

        # Check if final file already exists
        if final_path.exists():
            print(f"Skipped (already exists): {final_filename}")
            skipped += 1
            continue

        # Check if original file already exists
        if original_path.exists() and original_path != final_path:
            # File with original name exists, and we have a custom name
            if custom_name:
                # Rename the existing file
                try:
                    original_path.rename(final_path)
                    print(f"Renamed existing file: {original_filename} -> {final_filename}")
                    renamed += 1
                    skipped += 1
                    continue
                except Exception as e:
                    print(f"Error renaming {original_filename}: {str(e)}")
                    failed += 1
                    continue
            else:
                # No custom name, file already exists
                print(f"Skipped (already exists): {original_filename}")
                skipped += 1
                continue

        # Download the file to original path first
        if download_file(url, original_path):
            # If we have a custom name and it's different from original, rename
            if custom_name and final_path != original_path:
                try:
                    original_path.rename(final_path)
                    print(f"Downloaded and renamed: {final_filename}")
                    renamed += 1
                except Exception as e:
                    print(f"Downloaded as {original_filename}, but failed to rename: {str(e)}")
            else:
                print(f"Downloaded: {original_filename}")
            successful += 1
        else:
            failed += 1

    # Print summary
    print("\n" + "=" * 50)
    print("Download complete!")
    print(f"Successful: {successful}")
    print(f"Failed: {failed}")
    print(f"Skipped: {skipped}")
    if renamed > 0:
        print(f"Renamed: {renamed}")
    print(f"Total URLs: {len(unique_url_data)}")
    print("=" * 50)


if __name__ == "__main__":
    main()
